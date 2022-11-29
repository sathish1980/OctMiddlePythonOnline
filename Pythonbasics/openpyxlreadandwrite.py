import openpyxl


class openpyxlreadandwrite():

    inputfile="C:\\Users\\sathishkumar\\PycharmProjects\\OctMiddlePythonOnline\\input\\Testdata.xlsx"
    outputfile = "C:\\Users\\sathishkumar\\PycharmProjects\\OctMiddlePythonOnline\\output\\Testdataexcel.xlsx"
    def excelread(self):
        workbook=openpyxl.load_workbook(self.inputfile)
        sheetname=workbook["Input"]
        maxrow=sheetname.max_row
        print(maxrow)
        maxcolumn = sheetname.max_column
        print(maxcolumn)
        for eachrowvalue in range(1,maxrow+1): #row
            for eachcolumnvalue in range(1,maxcolumn+1): #column
                eachcellvalue=sheetname.cell(row=eachrowvalue, column=eachcolumnvalue).value # get the value
                print(eachcellvalue)

    def excelwrite(self):
        course=["python","java","sql","manual"]
        outputworkbook=openpyxl.Workbook()
        workbooksheet=outputworkbook.active
        for eachvalue in range(1,len(course)+1):
            #workbooksheet.cell(row=1, column=eachvalue).value=course[eachvalue-1]
            workbooksheet.cell(row=eachvalue, column=1).value = course[eachvalue - 1]
        outputworkbook.save(self.outputfile)
        print("done")

    def readwrite(self):
        workbook = openpyxl.load_workbook(self.inputfile)
        sheetname = workbook["Input"]
        #### output
        outputworkbook = openpyxl.Workbook()
        workbooksheet = outputworkbook.active

        maxrow = sheetname.max_row
        print(maxrow)
        maxcolumn = sheetname.max_column
        print(maxcolumn)
        for eachrowvalue in range(1, maxrow + 1):  # row
            for eachcolumnvalue in range(1, maxcolumn + 1):  # column
                eachcellvalue = sheetname.cell(row=eachrowvalue, column=eachcolumnvalue).value  # get the value
                print(eachcellvalue)
                workbooksheet.cell(row=eachrowvalue, column=eachcolumnvalue).value = eachcellvalue
        outputworkbook.save(self.outputfile)
        print("done")

obj=openpyxlreadandwrite()
obj.readwrite()

